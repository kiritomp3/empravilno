from __future__ import annotations

import tempfile
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path

import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook


@dataclass
class DaySummary:
    consumed_kcal: int
    burned_kcal: int
    net_kcal: int
    protein_g: float
    fat_g: float
    carb_g: float
    foods_count: int
    activities_count: int


def _to_decimal(v) -> Decimal:
    try:
        return Decimal(str(v))
    except (InvalidOperation, ValueError, TypeError):
        return Decimal("0")


def _round_macro(v, digits: int = 1) -> float:
    try:
        quant = Decimal("1") if digits == 0 else Decimal("1").scaleb(-digits)
        return float(_to_decimal(v).quantize(quant, rounding=ROUND_HALF_UP))
    except Exception:
        return 0.0


def _num(v, digits: int = 1) -> str:
    try:
        if isinstance(v, int):
            return str(v)
        f = float(v)
        s = f"{f:.{digits}f}".rstrip("0").rstrip(".")
        return s if s else "0"
    except Exception:
        return str(v)


def _build_summary(log: list[dict]) -> DaySummary:
    consumed = Decimal("0")
    burned = Decimal("0")
    protein = Decimal("0")
    fat = Decimal("0")
    carb = Decimal("0")
    foods_count = 0
    activities_count = 0

    for entry in log:
        entry_type = entry.get("entry_type", "food")
        total = entry.get("total") or {}
        if entry_type == "activity":
            burned += _to_decimal(total.get("burned_kcal", total.get("kcal", 0)))
            activities_count += 1
            continue

        consumed += _to_decimal(total.get("kcal", 0))
        protein += _to_decimal(total.get("protein_g", 0))
        fat += _to_decimal(total.get("fat_g", 0))
        carb += _to_decimal(total.get("carb_g", 0))
        foods_count += 1

    consumed_int = int(_round_macro(consumed, 0))
    burned_int = int(_round_macro(burned, 0))
    return DaySummary(
        consumed_kcal=consumed_int,
        burned_kcal=burned_int,
        net_kcal=consumed_int - burned_int,
        protein_g=_round_macro(protein, 1),
        fat_g=_round_macro(fat, 1),
        carb_g=_round_macro(carb, 1),
        foods_count=foods_count,
        activities_count=activities_count,
    )


def render_day_table(items: list[dict]) -> str:
    summary = _build_summary(items)
    return (
        "<pre>"
        f"Еда: {summary.foods_count} | Спорт: {summary.activities_count}\n"
        f"Калории: {summary.consumed_kcal} | Сожжено: {summary.burned_kcal} | Итог: {summary.net_kcal}\n"
        f"Б: {_num(summary.protein_g)} | Ж: {_num(summary.fat_g)} | У: {_num(summary.carb_g)}"
        "</pre>"
    )


def _food_rows(items: list[dict]) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for item in items:
        if item.get("entry_type", "food") != "food":
            continue
        total = item.get("total") or {}
        rows.append(
            {
                "Название": item.get("name", ""),
                "Кол-во": item.get("quantity", ""),
                "Ккал": int(_round_macro(total.get("kcal", 0), 0)),
                "Белки, г": _round_macro(total.get("protein_g", 0.0)),
                "Жиры, г": _round_macro(total.get("fat_g", 0.0)),
                "Углеводы, г": _round_macro(total.get("carb_g", 0.0)),
            }
        )
    return rows


def _activity_rows(items: list[dict]) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for item in items:
        if item.get("entry_type") != "activity":
            continue
        total = item.get("total") or {}
        details = item.get("details") or item.get("quantity") or ""
        rows.append(
            {
                "Активность": item.get("name", ""),
                "Детали": details,
                "Сожжено, ккал": int(
                    _round_macro(total.get("burned_kcal", total.get("kcal", 0)), 0)
                ),
            }
        )
    return rows


def _rows_to_df(rows: list[dict[str, object]], columns: list[str]) -> pd.DataFrame:
    return pd.DataFrame(rows, columns=columns)


def _style_table(table, n_rows: int, n_cols: int, header_color: str, accent_color: str) -> None:
    table.auto_set_font_size(False)
    table.set_fontsize(12)
    table.scale(1.15, 2.0)

    for (row, col), cell in table.get_celld().items():
        cell.set_edgecolor("#D7D4CC")
        cell.set_linewidth(0.8)
        if row == 0:
            cell.set_facecolor(header_color)
            cell.set_text_props(color="#FFFFFF", weight="bold")
        else:
            cell.set_facecolor("#FFFDF8" if row % 2 else "#F7F2E8")

    if n_rows > 0:
        for col in range(n_cols):
            table[(n_rows, col)].set_text_props(weight="bold", color=accent_color)


def save_dataframe_as_xlsx(
    foods_df: pd.DataFrame, activities_df: pd.DataFrame, base_name: str = "nutrition_day"
) -> Path:
    tmpdir = Path(tempfile.gettempdir())
    path = tmpdir / f"{base_name}.xlsx"
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        foods_df.to_excel(writer, index=False, sheet_name="День")
        startrow = len(foods_df) + 3
        activities_df.to_excel(writer, index=False, sheet_name="День", startrow=startrow)

    wb = load_workbook(path)
    ws = wb["День"]
    for row in ws.iter_rows(min_row=2, max_row=max(2, len(foods_df) + 1)):
        if len(row) >= 6:
            row[2].number_format = "0"
            row[3].number_format = "0.0"
            row[4].number_format = "0.0"
            row[5].number_format = "0.0"

    activity_start = len(foods_df) + 5
    for row in ws.iter_rows(min_row=activity_start + 1, max_row=activity_start + len(activities_df)):
        if len(row) >= 3:
            row[2].number_format = "0"

    wb.save(path)
    return path


def save_dataframe_as_csv(df: pd.DataFrame, base_name: str = "nutrition_day") -> Path:
    tmpdir = Path(tempfile.gettempdir())
    path = tmpdir / f"{base_name}.csv"
    df.to_csv(path, index=False)
    return path


def render_dataframe_to_png(
    foods_df: pd.DataFrame,
    activities_df: pd.DataFrame,
    base_name: str = "nutrition_day",
) -> Path:
    tmpdir = Path(tempfile.gettempdir())
    path = tmpdir / f"{base_name}.png"

    food_rows = max(len(foods_df), 1)
    activity_rows = max(len(activities_df), 1)
    fig_height = 2.8 + food_rows * 0.68 + activity_rows * 0.68

    fig = plt.figure(figsize=(14, fig_height), dpi=220, facecolor="#F6F0E5")
    fig.text(0.06, 0.955, "Дневник за день", fontsize=24, fontweight="bold", color="#2F3A2D")
    fig.text(0.06, 0.925, "Питание и спорт в одном отчёте", fontsize=11, color="#6D6A62")

    gs = fig.add_gridspec(2, 1, height_ratios=[food_rows + 1.5, activity_rows + 1.5], hspace=0.22)
    food_ax = fig.add_subplot(gs[0])
    act_ax = fig.add_subplot(gs[1])

    for ax in (food_ax, act_ax):
        ax.axis("off")
        ax.set_facecolor("#F6F0E5")

    food_ax.text(0.0, 1.08, "Питание", transform=food_ax.transAxes, fontsize=15, fontweight="bold", color="#2F3A2D")
    act_ax.text(0.0, 1.08, "Спорт активности", transform=act_ax.transAxes, fontsize=15, fontweight="bold", color="#2F3A2D")

    food_table = food_ax.table(
        cellText=foods_df.values.tolist() if not foods_df.empty else [["Пока нет записей", "", "", "", "", ""]],
        colLabels=list(foods_df.columns) if not foods_df.empty else ["Название", "Кол-во", "Ккал", "Белки, г", "Жиры, г", "Углеводы, г"],
        colWidths=[0.32, 0.16, 0.1, 0.14, 0.14, 0.14],
        loc="center",
    )
    _style_table(food_table, max(len(foods_df), 1), 6, "#497D4E", "#305233")

    act_table = act_ax.table(
        cellText=activities_df.values.tolist() if not activities_df.empty else [["Пока нет активностей", "", "0"]],
        colLabels=list(activities_df.columns) if not activities_df.empty else ["Активность", "Детали", "Сожжено, ккал"],
        colWidths=[0.34, 0.46, 0.2],
        loc="center",
    )
    _style_table(act_table, max(len(activities_df), 1), 3, "#D17A2B", "#8E4E15")

    plt.savefig(path, bbox_inches="tight", pad_inches=0.35, facecolor=fig.get_facecolor())
    plt.close(fig)
    return path


def build_day_files(items: list[dict], prefer_xlsx: bool = True) -> dict:
    foods_rows = _food_rows(items)
    activity_rows = _activity_rows(items)
    summary = _build_summary(items)

    foods_df = _rows_to_df(
        foods_rows + [
            {
                "Название": "ИТОГО",
                "Кол-во": "",
                "Ккал": summary.consumed_kcal,
                "Белки, г": summary.protein_g,
                "Жиры, г": summary.fat_g,
                "Углеводы, г": summary.carb_g,
            }
        ],
        ["Название", "Кол-во", "Ккал", "Белки, г", "Жиры, г", "Углеводы, г"],
    )
    activities_df = _rows_to_df(
        activity_rows + [{"Активность": "ИТОГО", "Детали": "", "Сожжено, ккал": summary.burned_kcal}],
        ["Активность", "Детали", "Сожжено, ккал"],
    )

    base = "nutrition_day"
    xlsx = save_dataframe_as_xlsx(foods_df, activities_df, base) if prefer_xlsx else None
    csv = None
    png = render_dataframe_to_png(foods_df, activities_df, base_name="nutrition_day_pretty")

    return {
        "png": png,
        "xlsx": xlsx,
        "csv": csv,
        "caption": "Дневник за день",
        "empty": not items,
        "summary": summary,
    }
